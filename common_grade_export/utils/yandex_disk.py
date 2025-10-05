import os
from os import environ, path

from openpyxl import load_workbook

import yadisk
from utils.gspread import add_csv_to_table


def write_sheet_to_file(yatoken, remote_path, csv_path, sheet_name="export"):
    disk_manager = DiskManagerMoodle(yatoken=yatoken)

    # download file to filesystem
    local_path = disk_manager.download_file_from_disk(remote_path)

    # create openpyxl.Workbook from existing xlsx file
    wb = load_workbook(filename=local_path)

    # add csv to table as sheet
    add_csv_to_table(csv_path, wb, sheet_name=sheet_name)

    # save openpyxl.Workbook to filesystem with same name
    wb.save(local_path)

    # download file to disk
    disk_manager.upload(local_path, remote_path)

    os.remove(local_path)


def write_sheet_to_file_dis(yatoken, remote_path, csv_path, sheet_name="export"):
    disk_manager = DiskManagerMoodle(yatoken=yatoken)

    # download file to filesystem
    local_path = disk_manager.download_file_from_disk(remote_path)

    # create openpyxl.Workbook from existing xlsx file
    wb = load_workbook(filename=local_path)

    # add csv to table as sheet
    add_csv_to_table(csv_path, wb, sheet_name=sheet_name)

    # save openpyxl.Workbook to filesystem with same name
    wb.save(local_path)

    # download file to disk
    disk_manager.upload(local_path, remote_path)

    os.remove(local_path)


class DiskManager:
    """Light YaDisk manager"""

    def __init__(self, yatoken=None, download_path="./"):
        self.client = yadisk.Client(token=yatoken or environ.get("YADISK_TOKEN"))
        self.download_path = download_path

    def upload(self, local_path: str, disk_path: str, overwrite=True):
        """upload from local_path to disk_path

        Args:
            local_path (str): path to local file
            disk_path (str): full path to file on yadisk
            overwrite (bool): overwrite file. Defaults to true
        """
        logger.info("Uploading %s to %s", *(local_path, disk_path))
        self.client.upload(local_path, disk_path, overwrite=overwrite)

    def download_file_from_disk(self, remote_path: str):
        """_summary_

        Args:
            remote_path (str): full path to file on yadisk

        Returns:
            str: path to downloaded file
        """
        local_path = self.download_path + path.basename(remote_path)
        self.client.download(remote_path, local_path)
        return local_path


class DiskManagerMoodle:
    """Light YaDisk manager"""

    def __init__(self, yatoken=None, download_path="./"):
        self.client = yadisk.Client(token=yatoken or environ.get("YADISK_TOKEN"))
        self.download_path = download_path

    def upload(self, local_path: str, disk_path: str, overwrite=True):
        """upload from local_path to disk_path

        Args:
            local_path (str): path to local file
            disk_path (str): full path to file on yadisk
            overwrite (bool): overwrite file. Defaults to true
        """
        logger.info("Uploading %s to %s", *(local_path, disk_path))
        self.client.upload(local_path, disk_path, overwrite=overwrite)

    def download_file_from_disk(self, remote_path: str):
        """_summary_

        Args:
            remote_path (str): full path to file on yadisk

        Returns:
            str: path to downloaded file
        """
        local_path = self.download_path + path.basename(remote_path)
        self.client.download(remote_path, local_path)
        return local_path
