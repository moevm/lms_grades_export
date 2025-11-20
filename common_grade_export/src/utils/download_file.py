import argparse
import gspread
from io import BytesIO
import logging
from openpyxl import load_workbook
import requests
from tempfile import NamedTemporaryFile
from pathlib import Path
from google.oauth2 import service_account
from google.auth.transport.requests import Request
from PyPDF2 import PdfMerger
from openpyxl import load_workbook, Workbook

logger = logging.getLogger(__name__)


# TODO: class for google service


def get_sheets_service_and_token(credentials_file="credentials.json"):
    """Создает и возвращает сервис для работы с Google Sheets и API access_token"""
    creds = service_account.Credentials.from_service_account_file(
        credentials_file,
        scopes=[
            "https://www.googleapis.com/auth/spreadsheets",
            "https://www.googleapis.com/auth/drive.file",
        ],
    )
    client = gspread.authorize(creds)
    creds.refresh(Request())
    return client, creds.token


def download_sheets(
    table_id: str,
    sheet_ids: list[str],
    filename: str = "export",
    export_format: str = "pdf",
    google_cred: str = "credentials.json",
    write_to_file: bool = True,
) -> bytes | None:
    """
    Скачивает несколько листов и объединяет их в один файл
    """
    try:
        client, access_token = get_sheets_service_and_token(google_cred)
        
        if len(sheet_ids) == 1:
            content = export_file(table_id, sheet_ids[0], access_token, export_format)
            if export_format == "xlsx" and content:
                content = get_excel_with_values(content)
        else:
            if export_format == "pdf":
                content = merge_multiple_pdfs(table_id, sheet_ids, access_token)
            elif export_format == "xlsx":
                content = merge_multiple_excels(table_id, sheet_ids, access_token)
            else:
                logger.warning(f"Формат {export_format} не поддерживает множественные листы, используется первый лист")
                content = export_file(table_id, sheet_ids[0], access_token, export_format)

        if not content:
            logger.error(f"Ошибка экспорта файла")
            return None

        if write_to_file:
            new_filepath = Path(f"{filename}.{export_format}")
            new_filepath.parents[0].mkdir(parents=True, exist_ok=True)
            with open(new_filepath, "wb") as f:
                f.write(content)
            logger.debug(f"Файл сохранен как: {new_filepath}")
        return content

    except Exception as e:
        logger.error(f"Ошибка при скачивании: {e}")
        return None


def merge_multiple_pdfs(table_id: str, sheet_ids: list[str], access_token: str) -> bytes:
    """
    Объединяет несколько PDF-файлов в один PDF-файл
    """
    merger = PdfMerger()
    temp_files = []
    
    try:
        for i, sheet_id in enumerate(sheet_ids):
            pdf_content = export_file(table_id, sheet_id, access_token, "pdf")
            if pdf_content:
                with NamedTemporaryFile(delete=False, suffix='.pdf') as temp_file:
                    temp_file.write(pdf_content)
                    temp_files.append(temp_file.name)
                    merger.append(temp_file.name)
        
        merged_pdf = BytesIO()
        merger.write(merged_pdf)
        merger.close()
        
        return merged_pdf.getvalue()
        
    finally:
        for temp_file in temp_files:
            try:
                Path(temp_file).unlink(missing_ok=True)
            except:
                pass


def merge_multiple_excels(table_id: str, sheet_ids: list[str], access_token: str) -> bytes:
    """
    Объединяет несколько листов в один XLSX-файл
    """
    merged_workbook = Workbook()    
    try:
        for i, sheet_id in enumerate(sheet_ids):
            excel_content = export_file(table_id, sheet_id, access_token, "xlsx")
            if excel_content:
                temp_wb = load_workbook(BytesIO(excel_content), data_only=True)
                
                for sheet_name in temp_wb.sheetnames:
                    source_sheet = temp_wb[sheet_name]
                    new_sheet = merged_workbook.create_sheet(title=f"{sheet_name}")
                    
                    for row in source_sheet.iter_rows():
                        for cell in row:
                            new_sheet[cell.coordinate].value = cell.value
        
        output = BytesIO()
        merged_workbook.save(output)
        output.seek(0)
        
        return output.getvalue()
        
    finally:
        merged_workbook.close()


def get_excel_with_values(content: bytes) -> bytes:
    """
    Сохраняет значения (не формулы) листа таблицы в XLSX-файл
    """
    wb = load_workbook(BytesIO(content), data_only=True)

    file_stream = BytesIO()
    wb.save(file_stream)
    file_stream.seek(0)

    return file_stream.read()


def export_file(
    table_id: str, sheet_id: str, access_token: str, export_format: str
) -> bytes | None:
    """
    Экспортирует файл используя export-url
    """
    url = f"https://docs.google.com/spreadsheets/d/{table_id}/export?format={export_format}&gid={sheet_id}"

    response = requests.get(url, headers={"Authorization": f"Bearer {access_token}"})

    if response.status_code == 200:
        return response.content
    else:
        logger.error(f"export_file: Ошибка {response.status_code}: {response.text}")
        return None

def parse_args():
    parser = argparse.ArgumentParser(description="Download Google Sheets")
    parser.add_argument("--table_id", required=True, help="Google Sheets table ID")
    parser.add_argument(
        "--sheet_ids", required=True, default="0", type=lambda x: x.split(";"), help="Sheet IDs separated by ; (default: 0)"
    )
    parser.add_argument(
        "--format", choices=["csv", "pdf", "xlsx"], default="csv", help="Output format"
    )
    parser.add_argument(
        "--filename", default="export", help="Output filename (without extension)"
    )
    parser.add_argument("--google_cred", help="Path to google credentials file")

    return parser.parse_args()


def main():
    args = parse_args()

    download_sheets(
        args.table_id, args.sheet_ids, args.filename, args.format, args.google_cred
    )


if __name__ == "__main__":
    main()
